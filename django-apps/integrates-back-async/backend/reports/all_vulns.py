# -*- coding: utf-8 -*-
""" Export all vulnerabilities """


import hashlib
import uuid
from typing import Dict, List, cast
from openpyxl import Workbook

from asgiref.sync import async_to_sync

from backend.dal import project as project_dal
from backend.domain import vulnerability as vuln_domain
from backend.typing import Finding as FindingType

from __init__ import FI_TEST_PROJECTS

TEST_PROJECTS = FI_TEST_PROJECTS.split(',')

COLUMNS_FINS = [
    'project_name',
    'finding_id',
    'finding',
    'finding_type',
    'attack_vector',
    'attack_complexity',
    'user_interaction',
    'severity_scope',
    'confidentiality_impact',
    'integrity_impact',
    'availability_impact',
    'exploitability',
    'remediation_level',
    'report_confidence',
    'cvss_basescore',
    'cvss_temporal',
    'actor',
    'cwe',
    'scenario'
]

COLUMNS_VULNS = [
    'vuln_type',
    'report_date',
    'analyst',
    'treatment',
    'specific',
    'closing_date'
]

MASKED_COLUMNS = [
    'finding_id',
    'project_name',
]


def _hash_cell(cell: str) -> str:
    return hashlib.sha256(cell.encode()).hexdigest()[-5:]


def _mask_finding(finding: Dict[str, FindingType]) -> Dict[str, FindingType]:
    for masked_cell in MASKED_COLUMNS:
        finding[masked_cell] = _hash_cell(str(finding[masked_cell]))
    return finding


def _get_reporter_analyst(
    opening_state: Dict[str, str],
    vuln: Dict[str, FindingType],
    finding: Dict[str, FindingType]
) -> str:
    analyst = opening_state.get('analyst', '')
    if not analyst:
        analyst = str(vuln.get('analyst', ''))
        if not analyst:
            finding.get('analyst', '')
    if analyst:
        analyst = _hash_cell(analyst)
    else:
        analyst = ''
    return analyst


def _format_specific(vuln: Dict[str, FindingType]) -> str:
    specific = ''
    if vuln.get('specific') != 'Masked':
        if vuln.get('vuln_type') == 'lines':
            file_ext = str(vuln.get('where')).split('/')[-1].split('.')[-1]
            if (file_ext.isalnum()
                    and not file_ext.isdigit()
                    and file_ext != 'Masked'):
                specific = file_ext.lower()
            else:
                specific = ''
        elif vuln.get('vuln_type') == 'ports':
            specific = str(vuln.get('specific'))
        else:
            specific = ''
    else:
        specific = ''

    return specific


def _format_vuln(
    vuln: Dict[str, FindingType],
    finding: Dict[str, FindingType]
) -> Dict[str, FindingType]:
    if not vuln.get('treatment'):
        if finding.get('TREATMENT'):
            vuln['treatment'] = finding.get('TREATMENT')
        else:
            vuln['treatment'] = 'NEW'

    historic_state = cast(List[Dict[str, str]], vuln.get('historic_state'))
    last_state = historic_state[-1]
    opening_state = historic_state[0]

    if last_state.get('state') == 'closed':
        vuln['treatment'] = 'CLOSED'
        vuln['closing_date'] = last_state.get('date')

    vuln['specific'] = _format_specific(vuln)
    vuln['report_date'] = opening_state.get('date')
    vuln['analyst'] = _get_reporter_analyst(opening_state, vuln, finding)
    return vuln


def fill_sheet(
    sheet,
    finding_row: Dict[str, FindingType],
    vuln_row: Dict[str, FindingType],
    row_index: int
):
    for col, col_name in enumerate(COLUMNS_FINS, 1):
        sheet.cell(row_index, col, finding_row.get(col_name, ''))
    for col, col_name in enumerate(COLUMNS_VULNS, len(COLUMNS_FINS) + 1):
        sheet.cell(row_index, col, vuln_row.get(col_name, ''))


def generate_all_vulns_xlsx(user_email: str, project_name: str = '') -> str:
    if project_name:
        projects = [{'project_name': project_name}]
    else:
        projects = cast(
            List[Dict[str, str]],
            project_dal.get_all(data_attr='project_name')
        )
    book = Workbook()
    sheet = book.active
    sheet.append(COLUMNS_FINS + COLUMNS_VULNS)
    row_index = 2
    for project in projects:
        if project not in TEST_PROJECTS:
            findings = async_to_sync(project_dal.get_released_findings)(
                project.get('project_name', ''))
        else:
            findings = []
        for finding in findings:
            vulns = async_to_sync(vuln_domain.list_vulnerabilities_async)(
                [str(finding['finding_id'])]
            )
            finding_row = cast(Dict[str, FindingType], _mask_finding(finding))
            for vuln in vulns:
                vuln_row = cast(
                    Dict[str, FindingType],
                    _format_vuln(vuln, finding_row)
                )
                fill_sheet(sheet, finding_row, vuln_row, row_index)
                row_index += 1

    username = user_email.split('@')[0]
    filepath = f'/tmp/{username}-{str(uuid.uuid4())}-allvulns.xlsx'
    book.save(filepath)
    return filepath
