# Standard library
from typing import cast, List
from openpyxl import load_workbook

# Local libraries
from backend.domain import (
    project as project_domain,
    vulnerability as vuln_domain
)
from backend.typing import Historic as HistoricType
from backend.utils import reports as reports_utils


# pylint: disable=too-many-locals
def generate(
    user_email: str,
    projects: List[str]
) -> str:
    template_path = (
        '/usr/src/app/django-apps/integrates-back-async/backend'
        '/reports/templates/excel/COMPLETE.xlsx'
    )
    book = load_workbook(template_path)
    sheet = book.active

    project_col = 1
    finding_col = 2
    vuln_where_col = 3
    vuln_specific_col = 4
    treatment_col = 5
    treatment_mgr_col = 6
    row_offset = 2

    row_index = row_offset
    for project in projects:
        findings = project_domain.get_released_findings(
            project, 'finding_id, finding, historic_treatment'
        )
        for finding in findings:
            vulns = vuln_domain.get_vulnerabilities(str(finding['finding_id']))
            for vuln in vulns:
                sheet.cell(row_index, vuln_where_col, vuln['where'])
                sheet.cell(row_index, vuln_specific_col, vuln['specific'])

                sheet.cell(row_index, project_col, project.upper())
                sheet.cell(
                    row_index,
                    finding_col,
                    (f'{str(finding["finding"]).encode("utf-8")!s} '
                     f'(#{str(finding["finding_id"])})')
                )
                historic_treatment = finding.get('historic_treatment', [{}])
                sheet.cell(
                    row_index,
                    treatment_col,
                    cast(
                        HistoricType,
                        historic_treatment
                    )[-1].get('treatment', '')
                )
                sheet.cell(
                    row_index,
                    treatment_mgr_col,
                    vuln.get('treatment_manager', 'Unassigned')
                )

                row_index += 1

    username = user_email.split('@')[0]
    report_filename = 'complete_report.xlsx'
    report_filepath = f'/tmp/{username}-{report_filename}'
    book.save(cast(str, report_filepath))
    uploaded_file_name = reports_utils.upload_report(report_filepath)
    uploaded_file_url = reports_utils.sign_url(
        uploaded_file_name,
        minutes=1.0 / 6
    )

    return uploaded_file_url
