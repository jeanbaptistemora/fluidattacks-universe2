from typing import cast, Dict
import uuid
from openpyxl import Workbook
from backend.dal import user as user_dal
from backend.domain import user as user_domain
from backend.utils import reports as reports_utils


def generate(user_email: str) -> str:
    book = Workbook()
    sheet = book.active
    sheet.append(['full_name', 'user_email'])
    row_index = 2

    unique_users = []
    for user in user_dal.get_platform_users():
        email = str(user.get('user_email', '')).lower()
        if email not in unique_users:
            unique_users.append(email)

            name_attrs = cast(
                Dict[str, str],
                user_domain.get_attributes(email, ['first_name', 'last_name'])
            )
            full_name = ' '.join(list(name_attrs.values()))

            sheet.cell(row_index, 1, full_name)
            sheet.cell(row_index, 2, email)
            row_index += 1

    username = user_email.split('@')[0]
    report_filepath = f'/tmp/{username}-{uuid.uuid4()}-allusers.xlsx'
    book.save(report_filepath)

    uploaded_file_name = reports_utils.upload_report(report_filepath)
    uploaded_file_url = reports_utils.sign_url(
        uploaded_file_name,
        minutes=1.0 / 6
    )

    return uploaded_file_url
