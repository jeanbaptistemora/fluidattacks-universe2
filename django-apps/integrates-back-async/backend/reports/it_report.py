# -*- coding: utf-8 -*-
""" Class for generate an xlsx file with findings information. """
import os
from typing import cast, Dict, List, Optional, Union

from datetime import datetime
from django.conf import settings
from asgiref.sync import async_to_sync
import pytz
from openpyxl import load_workbook
from pyexcelerate import (
    Color,
    Style,
    Workbook,
    Worksheet as WorksheetType
)
from backend.domain import vulnerability as vuln_domain
from backend.domain.finding import get_finding
from backend.typing import (
    Finding as FindingType,
    Historic as HistoricType,
    Vulnerability as VulnType
)


EMPTY = '-'
HEADER_HEIGHT = 20
ROW_HEIGHT = 50
RED = Color(255, 52, 53, 1)  # FF3435
WHITE = Color(255, 255, 255, 1)


class ITReport():
    """Class to generate IT reports."""

    workbook = None
    template = None
    current_sheet: WorksheetType = None
    data = None
    lang = None
    row = 1
    result_filename = ''
    project_name = ''
    base = (
        '/usr/src/app/django-apps/integrates-back-async/backend/reports'
    )
    result_path = os.path.join(base, 'results/results_excel')
    templates: Dict[str, Dict[str, str]] = {
        'es': {
            'TECHNICAL': os.path.join(
                base, 'templates/excel', 'TECHNICAL_VULNS.xlsx'),
        },
        'en': {}}
    sheet_names = {
        'es': {
            'data': 'Data',
        },
        'en': {}}
    vulnerability = {
        'number': 1,
        'finding': 2,
        'finding_id': 3,
        'vuln_uuid': 4,
        'where': 5,
        'specific': 6,
        'description': 7,
        'status': 8,
        'severity': 9,
        'requirements': 10,
        'impact': 11,
        'affected_systems': 12,
        'threat': 13,
        'recommendation': 14,
        'external_bts': 15,
        'compromised_attributes': 16,
        'n_compromised_attributes': 17,
        'tags': 18,
        'business_critically': 19,
        'vuln_report_date': 20,
        'vuln_close_date': 21,
        'vuln_age': 22,
        'first_treatment': 23,
        'first_treatment_date': 24,
        'first_treatment_justification': 25,
        'first_treatment_exp_date': 26,
        'first_treatment_manager': 27,
        'treatment': 28,
        'treatment_date': 29,
        'treatment_justification': 30,
        'treatment_exp_date': 31,
        'treatment_manager': 32,
        'reattack': 33,
        'n_requested_reattacks': 34,
        'remediation_effectiveness': 35,
        'last_reattack_date': 36,
        'last_reattack_requester': 37,
        'cvss_vector': 38,
        'AV': 39, 'AC': 40, 'PR': 41, 'UI': 42,
        'SS': 43, 'CI': 44, 'II': 45, 'AI': 46,
        'E': 47, 'RL': 48, 'RC': 49,
        '': 50
    }
    align_left = [2, 7, 10, 11, 12, 13, 14, 15, 16, 18, 25, 30]
    row_values: List[Union[str, int, datetime]] = \
        [EMPTY for _ in vulnerability]

    def __init__(self, data: List[Dict[str, FindingType]], lang: str = 'es'):
        """Initialize variables."""
        self.project_name = str(data[0].get('projectName'))
        self.lang = lang
        self.template = load_workbook(
            filename=self.templates[self.lang]['TECHNICAL']
        )
        self.workbook = Workbook()
        self.current_sheet = self.workbook.new_sheet('Data')

        self.parse_template()
        self.generate(data)
        self.style_sheet()

        self.save()

    def save(self):
        tzn = pytz.timezone(settings.TIME_ZONE)  # type: ignore
        today_date = datetime.now(tz=tzn).today().strftime('%Y-%m-%dT%H-%M-%S')
        self.result_filename = (
            f'{self.project_name}-vulnerabilities-{today_date}.xlsx'
        )
        self.workbook.save(self.result_filename)

    def style_sheet(self):
        header = self.current_sheet.range(*self.get_row_range(1))
        header.style.fill.background = RED
        header.style.font.color = WHITE
        header.style.alignment.horizontal = 'center'
        header.style.alignment.vertical = 'center'
        header.style.alignment.wrap_text = True

        for column in range(1, len(self.vulnerability.values())):
            self.current_sheet.set_col_style(
                column,
                Style(size=-1)
            )
        self.current_sheet.set_row_style(2, Style(size=ROW_HEIGHT))

    def parse_template(self):
        template_sheet = self.template[
            self.sheet_names[self.lang]['data']
        ]
        self.current_sheet.range(*self.get_row_range(self.row)).value = [
            [str(template_sheet.cell(row=self.row, column=column).value)
             for column in range(1, len(self.vulnerability.values()))]
        ]
        self.row += 1

    def generate(self, data: List[Dict[str, FindingType]]):
        self.project_name = str(data[0].get('projectName'))
        vulns = async_to_sync(vuln_domain.list_vulnerabilities_async)(
            [finding.get('findingId') for finding in data])
        for vuln in vulns:
            self.write_vuln_row(vuln)
            self.row += 1

    @classmethod
    def get_row_range(cls, row: int) -> List[str]:
        return [f'A{row}', f'AW{row}']

    @classmethod
    def get_measure(
        cls,
        metric: str,
        metric_value: str
    ) -> Optional[str]: # noqa
        """Extract number of CSSV metrics."""
        try:
            metrics = {
                'attackVector': {
                    '0.85': 'Network',
                    '0.62': 'Adjacent',
                    '0.55': 'Local',
                    '0.20': 'Physical',
                },
                'attackComplexity': {
                    '0.77': 'Low',
                    '0.44': 'High',
                },
                'privilegesRequired': {
                    '0.85': 'None',
                    '0.62': 'Low',
                    '0.68': 'Low',
                    '0.27': 'High',
                    '0.50': 'High',
                },
                'userInteraction': {
                    '0.85': 'None',
                    '0.62': 'Required',
                },
                'severityScope': {
                    '0.0': 'Unchanged',
                    '1.0': 'Changed',
                },
                'confidentialityImpact': {
                    '0.56': 'High',
                    '0.22': 'Low',
                    '0.0': 'None',
                },
                'integrityImpact': {
                    '0.56': 'High',
                    '0.22': 'Low',
                    '0.0': 'None',
                },
                'availabilityImpact': {
                    '0.56': 'High',
                    '0.22': 'Low',
                    '0.0': 'None',
                },
                'exploitability': {
                    '0.91': 'Unproven',
                    '0.94': 'Proof of concept',
                    '0.97': 'Functional',
                    '1.0': 'High',
                },
                'remediationLevel': {
                    '0.95': 'Official Fix',
                    '0.96': 'Temporary Fix',
                    '0.97': 'Workaround',
                    '1.0': 'Unavailable',
                },
                'reportConfidence': {
                    '0.92': 'Unknown',
                    '0.96': 'Reasonable',
                    '1.0': 'Confirmed',
                }
            }
            metric_descriptions = metrics.get(metric)
            if metric_descriptions:
                description = metric_descriptions.get(str(metric_value))
            else:
                description = ''
            return description
        except ValueError:
            return ''

    def set_cvss_metrics_cell(self, row: VulnType):
        measures = {
            'AV': 'attackVector',
            'AC': 'attackComplexity',
            'PR': 'privilegesRequired',
            'UI': 'userInteraction',
            'S': 'severityScope',
            'C': 'confidentialityImpact',
            'I': 'integrityImpact',
            'A': 'availabilityImpact',
            'E': 'exploitability',
            'RL': 'remediationLevel',
            'RC': 'reportConfidence',
        }
        metric_vector = []
        for index, (indicator, measure) in enumerate(measures.items()):
            value = self.get_measure(
                measure,
                cast(Dict[str, str], row['severity'])[measure]
            )
            if value:
                metric_vector.append(f'{indicator}:{value[0]}')
                self.row_values[
                    self.vulnerability['cvss_vector'] + index + 1] = value

        cvss_metric_vector = '/'.join(metric_vector)
        cvss_calculator_url = (
            f'https://www.first.org/cvss/calculator/3.1#CVSS:3.1'
            f'/{cvss_metric_vector}'
        )
        cell_content = \
            f'=HYPERLINK("{cvss_calculator_url}", "{cvss_metric_vector}")'
        self.row_values[self.vulnerability['cvss_vector']] = cell_content

    def write_vuln_row(self, row: VulnType):
        finding = async_to_sync(get_finding)(row.get('finding_id'))
        specific = str(row.get('specific', ''))
        if row.get('vuln_type') == 'lines':
            specific = str(int(specific))
        tags = EMPTY
        if 'tag' in row:
            tags = str(', '.join(cast(List[str], row.get('tag'))))

        self.row_values[self.vulnerability['number']] = self.row - 1
        self.row_values[self.vulnerability['finding']] = finding.get('finding')
        self.row_values[self.vulnerability['finding_id']] = \
            finding.get('findingId', EMPTY)
        self.row_values[self.vulnerability['vuln_uuid']] = \
            str(row.get('UUID', EMPTY))
        self.row_values[self.vulnerability['where']] = str(row.get('where'))
        self.row_values[self.vulnerability['specific']] = specific
        self.row_values[self.vulnerability['tags']] = tags

        self.write_finding_data(finding, row)
        self.write_vuln_temporal_data(row)
        self.write_treatment_data(finding, row)
        self.write_reattack_data(finding, row)
        self.set_cvss_metrics_cell(finding)

        self.current_sheet.range(*self.get_row_range(self.row)).value = \
            [self.row_values[1:]]
        self.current_sheet.set_row_style(self.row, Style(size=ROW_HEIGHT))

    def write_finding_data(
        self,
        finding: Dict[str, FindingType],
        vuln: VulnType
    ):
        compromised_attributes = str(finding.get('compromisedAttrs')) or EMPTY
        n_compromised_attributes = None
        if compromised_attributes != EMPTY:
            n_compromised_attributes = \
                str(len(compromised_attributes.split('\n')))
        external_bts = finding.get('externalBts', EMPTY)

        finding_data = {
            'description': str(finding.get('vulnerability', EMPTY)),
            'status': cast(
                HistoricType, vuln.get('historic_state'))[-1]['state'],
            'severity': str(finding.get('severityCvss', EMPTY)),
            'requirements': str(finding.get('requirements', EMPTY)),
            'impact': str(finding.get('attackVectorDesc', EMPTY)),
            'affected_systems': str(finding.get('affectedSystems', EMPTY)),
            'threat': str(finding.get('threat', EMPTY)),
            'recommendation': str(finding.get('effectSolution', EMPTY)),
            'external_bts': f'=HYPERLINK("{external_bts}", "{external_bts}")',
            'compromised_attributes': compromised_attributes,
            'n_compromised_attributes': n_compromised_attributes or '0',
        }
        for key, value in finding_data.items():
            self.row_values[self.vulnerability[key]] = value

    def write_vuln_temporal_data(self, vuln: VulnType):
        vuln_historic_state = cast(HistoricType, vuln.get('historic_state'))
        vuln_date = datetime.strptime(
            vuln_historic_state[0]['date'], '%Y-%m-%d %H:%M:%S')
        vuln_closed = vuln_historic_state[-1]['state'] == 'closed'
        limit_date = datetime.today()
        vuln_close_date: Union[str, datetime] = EMPTY
        if vuln_closed:
            limit_date = datetime.strptime(
                vuln_historic_state[-1]['date'], '%Y-%m-%d %H:%M:%S'
            )
            vuln_close_date = datetime.strptime(
                vuln_historic_state[-1]['date'], '%Y-%m-%d %H:%M:%S'
            )
        vuln_age_days = int((limit_date - vuln_date).days)
        vuln_age = f'{vuln_age_days} '

        vuln_temporal_data: Dict[str, Union[str, int, datetime]] = {
            'vuln_report_date': vuln_date,
            'vuln_age': vuln_age,
            'vuln_close_date': vuln_close_date
        }
        for key, value in vuln_temporal_data.items():
            self.row_values[self.vulnerability[key]] = value

    def write_treatment_data(  # pylint: disable=too-many-locals
        self,
        finding: Dict[str, FindingType],
        vuln: VulnType
    ):
        def format_treatment(treatment: str) -> str:
            treatment = treatment.capitalize().replace('_', ' ')
            if treatment == 'Accepted undefined':
                treatment = 'Eternally accepted'
            elif treatment == 'Accepted':
                treatment = 'Temporarily accepted'
            return treatment

        historic_state = finding.get('historicState')
        finding_historic_treatment = \
            cast(HistoricType, finding.get('historicTreatment'))
        current_treatment_date: Union[str, datetime] = EMPTY
        current_treatment_exp_date: Union[str, datetime] = EMPTY
        first_treatment_exp_date: Union[str, datetime] = EMPTY
        if historic_state and 'date' in cast(HistoricType, historic_state)[-1]:
            current_treatment_date = datetime.strptime(
                str(cast(HistoricType, historic_state)[-1]['date']),
                '%Y-%m-%d %H:%M:%S'
            )
        if 'acceptance_date' in vuln:
            current_treatment_exp_date = datetime.strptime(
                str(vuln.get('acceptance_date')),
                '%Y-%m-%d %H:%M:%S'
            )
        first_treatment_state = \
            cast(Dict[str, str], finding_historic_treatment[0])
        if len(str(finding.get('releaseDate')).split(' ')) == 2:
            first_treatment_date_format = '%Y-%m-%d %H:%M:%S'
        else:
            first_treatment_date_format = '%Y-%m-%d'
        if len(finding_historic_treatment) > 1:
            first_treatment_exp_date = \
                finding_historic_treatment[1].get('date', EMPTY)
            if first_treatment_exp_date != EMPTY:
                first_treatment_exp_date = datetime.strptime(
                    str(first_treatment_exp_date),
                    '%Y-%m-%d %H:%M:%S'
                )

        current_treatment_data: Dict[str, Union[str, int, datetime]] = {
            'treatment': format_treatment(str(vuln.get('treatment', 'NEW'))),
            'treatment_date': current_treatment_date,
            'treatment_justification': str(
                vuln.get('treatment_justification', EMPTY)),
            'treatment_exp_date': current_treatment_exp_date,
            'treatment_manager': str(vuln.get('treatment_manager', EMPTY)),
        }
        first_treatment_data: Dict[str, Union[str, int, datetime]] = {
            'first_treatment': str(format_treatment(
                first_treatment_state.get('treatment', 'NEW'))),
            'first_treatment_date': datetime.strptime(
                str(finding.get('releaseDate')), first_treatment_date_format),
            'first_treatment_justification': str(
                first_treatment_state.get('justification', EMPTY)),
            'first_treatment_exp_date': first_treatment_exp_date,
            'first_treatment_manager': str(
                first_treatment_state.get('user', EMPTY)),
        }

        for key, value in current_treatment_data.items():
            self.row_values[self.vulnerability[key]] = value
            self.row_values[self.vulnerability[f'first_{key}']] = \
                first_treatment_data[f'first_{key}']

    def write_reattack_data(
        self,
        finding: Dict[str, FindingType],
        vuln: VulnType
    ):
        historic_verification = \
            cast(HistoricType, finding.get('historicVerification'))
        vuln_closed = cast(
            HistoricType, vuln.get('historic_state'))[-1]['state'] == 'closed'
        reattack_requested = None
        reattack_date = None
        reattack_requester = None
        n_requested_reattacks = None
        remediation_effectiveness = EMPTY
        if historic_verification:
            reattack_requested = \
                historic_verification[-1]['status'] == 'REQUESTED'
            n_requested_reattacks = \
                len([
                    state for state in historic_verification
                    if state['status'] == 'REQUESTED'
                ])
            if vuln_closed:
                remediation_effectiveness = f'{100 / n_requested_reattacks}%'
            if reattack_requested:
                reattack_date = datetime.strptime(
                    historic_verification[-1]['date'], '%Y-%m-%d %H:%M:%S')
                reattack_requester = historic_verification[-1]['user']
        reattack_data = {
            'reattack': 'Yes' if reattack_requested else 'No',
            'n_requested_reattacks': n_requested_reattacks or '0',
            'last_reattack_date': reattack_date or EMPTY,
            'last_reattack_requester': reattack_requester or EMPTY,
            'remediation_effectiveness': remediation_effectiveness
        }
        for key, value in reattack_data.items():
            self.row_values[self.vulnerability[key]] = value
