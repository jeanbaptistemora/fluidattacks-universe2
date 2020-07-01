# -*- coding: utf-8 -*-
# Disabling this rule is necessary for include returns inside if-else structure
# pylint: disable-msg=no-else-return
# pylint: disable=too-many-lines
"""Views and services for FluidIntegrates."""

# Standard library
import os
import sys
from datetime import datetime, timedelta
from typing import (
    List,
)

# Third party libraries
import boto3
import rollbar
from asgiref.sync import async_to_sync
from django.conf import settings
from django.core.cache.backends.base import DEFAULT_TIMEOUT
from django.http import HttpRequest, HttpResponse
from django.shortcuts import render, redirect
from django.views.decorators.cache import never_cache, cache_control
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from jose import jwt
from magic import Magic
from openpyxl import Workbook

# Local libraries
from backend import authz, util
from backend.domain import (
    analytics as analytics_domain,
    user as user_domain,
)
from backend.decorators import (
    authenticate,
    cache_content,
    require_login,
)
from backend.dal import (
    user as user_dal
)
from backend.exceptions import ConcurrentSession
from backend.services import (
    has_access_to_finding, has_access_to_event
)
from __init__ import (
    FI_AWS_S3_ACCESS_KEY, FI_AWS_S3_SECRET_KEY, FI_AWS_S3_BUCKET, FI_ENVIRONMENT
)
from app.documentator.all_vulns import generate_all_vulns_xlsx

# Constants
CACHE_TTL = getattr(settings, 'CACHE_TTL', DEFAULT_TIMEOUT)

CLIENT_S3 = boto3.client('s3',
                         aws_access_key_id=FI_AWS_S3_ACCESS_KEY,
                         aws_secret_access_key=FI_AWS_S3_SECRET_KEY,
                         aws_session_token=os.environ.get('AWS_SESSION_TOKEN'))

BUCKET_S3 = FI_AWS_S3_BUCKET


def enforce_user_level_role(request, *allowed_roles):
    # Verify role if the user is logged in
    email = request.session.get('username')
    registered = request.session.get('registered')

    if not email or not registered:
        # The user is not even authenticated. Redirect to login
        return HttpResponse("""
            <script>
                var getUrl=window.location.href.split(`${window.location.host}/integrates`);
                localStorage.setItem("start_url",getUrl[getUrl.length - 1]);
                location = "/integrates/index";
            </script>
            """)

    requester_role = authz.get_user_level_role(email)
    if requester_role not in allowed_roles:
        response = HttpResponse("Access denied")
        response.status_code = 403
        return response

    return None


def enforce_group_level_role(request, group, *allowed_roles):
    # Verify role if the user is logged in
    email = request.session.get('username')
    registered = request.session.get('registered')

    if not email or not registered:
        # The user is not even authenticated. Redirect to login
        return HttpResponse("""
            <script>
                var getUrl=window.location.href.split(`${window.location.host}/integrates`);
                localStorage.setItem("start_url",getUrl[getUrl.length - 1]);
                location = "/integrates/index";
            </script>
            """)

    requester_role = authz.get_group_level_role(email, group)
    if requester_role not in allowed_roles:
        response = HttpResponse("Access denied")
        response.status_code = 403
        return response

    return None


@never_cache
def index(request):
    """Login view for unauthenticated users"""
    parameters = {'debug': settings.DEBUG}
    return render(request, 'index.html', parameters)


def error500(request):
    """Internal server error view"""
    parameters = {}
    return render(request, 'HTTP500.html', parameters)


def error401(request, _):
    """Unauthorized error view"""
    parameters = {}
    return render(request, 'HTTP401.html', parameters)


@csrf_exempt
@cache_control(private=True, max_age=3600)
@authenticate
def app(request):
    """App view for authenticated users."""
    try:
        if FI_ENVIRONMENT == 'production':
            util.check_concurrent_sessions(
                request.session['username'], request.session.session_key)
        parameters = {
            'debug': settings.DEBUG,
            'username': request.session['username']
        }
        response = render(request, 'app.html', parameters)
        payload = {
            'user_email': request.session['username'],
            'company': request.session['company'],
            'first_name': request.session['first_name'],
            'last_name': request.session['last_name'],
            'exp': datetime.utcnow() +
            timedelta(seconds=settings.SESSION_COOKIE_AGE),
            'sub': 'django_session',
            'jti': util.calculate_hash_token()['jti'],
        }
        token = jwt.encode(
            payload,
            algorithm='HS512',
            key=settings.JWT_SECRET,
        )
        jti = payload['jti']
        util.save_token(f'fi_jwt:{jti}', token, settings.SESSION_COOKIE_AGE)
        response.set_cookie(
            key=settings.JWT_COOKIE_NAME,
            samesite=settings.JWT_COOKIE_SAMESITE,
            value=token,
            secure=True,
            # Temporary while ariadne migration is finished
            httponly=not settings.DEBUG,
            max_age=settings.SESSION_COOKIE_AGE
        )
    except KeyError:
        rollbar.report_exc_info(sys.exc_info(), request)
        return redirect('/integrates/error500')
    except ConcurrentSession:
        return HttpResponse("""
            <script>
                localStorage.setItem("concurrentSession","1");
                location.assign("/integrates/registration");
            </script>
            """)
    return response


@csrf_exempt
@require_login
@require_http_methods(['GET'])
@async_to_sync
async def graphic(request: HttpRequest) -> HttpResponse:
    return await analytics_domain.handle_graphic_request(request)


@csrf_exempt
@authenticate
def logout(request):
    """Close a user's active session"""
    try:
        cookie_content = jwt.decode(token=request.COOKIES.get(settings.JWT_COOKIE_NAME),
                                    key=settings.JWT_SECRET,
                                    algorithms='HS512')
        jti = cookie_content.get('jti')
        if jti:
            util.remove_token(f'fi_jwt:{jti}')

        request.session.flush()
    except KeyError:
        rollbar.report_exc_info(sys.exc_info(), request)

    response = redirect('/integrates/index')
    response.delete_cookie(settings.JWT_COOKIE_NAME)
    return response


@cache_content
@cache_control(private=True, max_age=31536000)
@csrf_exempt
def get_evidence(request, project, evidence_type, findingid, fileid):
    allowed_roles = [
        'admin', 'analyst', 'closer', 'customer', 'customeradmin', 'executive',
        'group_manager', 'internal_manager', 'resourcer', 'reviewer'
    ]

    error = enforce_group_level_role(request, project, *allowed_roles)

    if error is not None:
        return error

    username = request.session['username']
    if (evidence_type in ['drafts', 'findings']
        and has_access_to_finding(username, findingid)) \
            or (evidence_type == 'events'
                and has_access_to_event(username, findingid)):
        if fileid is None:
            rollbar.report_message('Error: Missing evidence image ID',
                                   'error', request)
            return HttpResponse('Error - Unsent image ID',
                                content_type='text/html')
        key_list = list_s3_evidences(f'{project.lower()}/{findingid}/{fileid}')
        if key_list:
            for k in key_list:
                start = k.find(findingid) + len(findingid)
                localfile = '/tmp' + k[start:]
                ext = {'.png': '.tmp', '.gif': '.tmp'}
                localtmp = util.replace_all(localfile, ext)
                CLIENT_S3.download_file(BUCKET_S3, k, localtmp)
                return retrieve_image(request, localtmp)
        else:
            return util.response([], 'Access denied or evidence not found', True)
    else:
        util.cloudwatch_log(
            request,
            'Security: Attempted to retrieve evidence without permission')
        return util.response([], 'Access denied or evidence not found', True)


def retrieve_image(request, img_file):
    if util.assert_file_mime(img_file, ['image/png', 'image/jpeg',
                                        'image/gif']):
        with open(img_file, 'rb') as file_obj:
            mime = Magic(mime=True)
            mime_type = mime.from_file(img_file)
            return HttpResponse(file_obj.read(), content_type=mime_type)
    else:
        rollbar.report_message('Error: Invalid evidence image format',
                               'error', request)
        return HttpResponse('Error: Invalid evidence image format',
                            content_type='text/html')


def list_s3_evidences(prefix) -> List[str]:
    """return keys that begin with prefix from the evidences folder."""
    return list(util.iterate_s3_keys(CLIENT_S3, BUCKET_S3, prefix))


@cache_content
@never_cache
def export_all_vulnerabilities(request):
    allowed_roles = ['admin']

    error = enforce_user_level_role(request, *allowed_roles)

    if error is not None:
        return error

    user_data = util.get_jwt_content(request)
    filepath = generate_all_vulns_xlsx(user_data['user_email'])
    filename = os.path.basename(filepath)
    with open(filepath, 'rb') as document:
        response = HttpResponse(document.read())
        response['Content-Type'] = 'application/vnd.openxmlformats\
                        -officedocument.spreadsheetml.sheet'
        response['Content-Disposition'] = 'inline;filename={filename}'.format(
            filename=filename)
    return response


@cache_content
@never_cache
def export_users(request):  # pylint: disable=too-many-locals
    allowed_roles = ['admin']

    error = enforce_user_level_role(request, *allowed_roles)

    if error is not None:
        return error

    user_data = util.get_jwt_content(request)
    book = Workbook()
    sheet = book.active
    sheet.append(['full_name', 'user_email'])
    row_index = 2

    unique_users = []
    for user in user_dal.get_platform_users():
        user_email = user['user_email'].lower()
        if user_email not in unique_users:
            unique_users.append(user_email)

            name_attrs = user_domain.get_attributes(
                user_email, ['first_name', 'last_name'])
            full_name = ' '.join(list(name_attrs.values()))

            sheet.cell(row_index, 1, full_name)
            sheet.cell(row_index, 2, user_email)
            row_index += 1

    username = user_data['user_email'].split('@')[0].encode('utf8', 'ignore')
    filepath = f'/tmp/{username}-users.xlsx'
    filename = os.path.basename(filepath)
    book.save(filepath)

    with open(filepath, 'rb') as document:
        response = HttpResponse(document.read())
        response['Content-Type'] = 'application/vnd.openxmlformats\
                        -officedocument.spreadsheetml.sheet'
        response['Content-Disposition'] = f'inline;filename={filename}'
    return response
